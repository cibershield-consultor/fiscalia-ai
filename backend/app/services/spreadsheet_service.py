"""
FiscalIA — Servicio de hojas de cálculo con IA inteligente
"""
import io, csv, json
from datetime import datetime
from typing import Optional
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter

# ── Paleta corporativa ─────────────────────────────────────────
NAVY   = "0D1B2A"
GOLD   = "C9A84C"
WHITE  = "FFFFFF"
GREEN  = "1A7A4A"
RED_   = "C0392B"
GREY_L = "F5F5F5"
GREY_D = "CCCCCC"
BLUE_  = "1F4E79"
LBLUE  = "D6E4F0"
LGREEN = "E8F5E9"
LORANG = "FFF3E0"

def _hfill(c): return PatternFill("solid", start_color=c, fgColor=c)
def _font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)
def _border():
    s = Side(style="thin", color=GREY_D)
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _right():  return Alignment(horizontal="right",  vertical="center")
def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

EUR_FMT  = '#,##0.00 €;(#,##0.00 €);"-"'
PCT_FMT  = '0.0%;(0.0%);"-"'
DATE_FMT = 'DD/MM/YYYY'


# ══════════════════════════════════════════════════════════════
#  LEER — parse Excel/CSV → text for AI
# ══════════════════════════════════════════════════════════════

def parse_spreadsheet_for_ai(file_bytes: bytes, filename: str) -> str:
    ext = filename.lower().split(".")[-1]
    if ext == "csv":
        return _parse_csv(file_bytes)
    elif ext in ("xlsx", "xls", "xlsm"):
        return _parse_excel(file_bytes)
    raise ValueError(f"Formato no soportado: {ext}")


def _parse_csv(file_bytes: bytes) -> str:
    try:   text = file_bytes.decode("utf-8")
    except: text = file_bytes.decode("latin-1")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows: return "CSV vacío."
    headers = rows[0]
    data_rows = rows[1:]
    lines = [f"CSV — {len(data_rows)} filas, {len(headers)} columnas",
             f"Columnas: {' | '.join(headers)}", ""]
    for col_idx, col_name in enumerate(headers):
        vals = []
        for row in data_rows:
            if col_idx < len(row):
                try: vals.append(float(row[col_idx].replace("€","").replace(",",".").strip()))
                except: pass
        if vals:
            lines.append(f"{col_name}: total={sum(vals):.2f}, avg={sum(vals)/len(vals):.2f}, min={min(vals):.2f}, max={max(vals):.2f}")
    lines += ["", "Primeras 20 filas:"]
    for i, row in enumerate(data_rows[:20]):
        lines.append(f"  Fila {i+1}: " + " | ".join(f"{headers[j] if j<len(headers) else j}={v}" for j,v in enumerate(row)))
    if len(data_rows) > 20:
        lines.append(f"  ... y {len(data_rows)-20} más")
    return "\n".join(lines)


def _parse_excel(file_bytes: bytes) -> str:
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    lines = [f"Excel — {len(wb.sheetnames)} hoja(s): {', '.join(wb.sheetnames)}"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n== HOJA: {sheet_name} ==")
        data = []
        for row in ws.iter_rows(values_only=True):
            if any(c is not None for c in row):
                data.append([str(c) if c is not None else "" for c in row])
        if not data: lines.append("  (vacía)"); continue
        headers = data[0]
        lines.append(f"  Dimensiones: {ws.max_row}×{ws.max_column}")
        lines.append(f"  Columnas: {' | '.join(h for h in headers if h)}")
        for col_idx, col_name in enumerate(headers):
            if not col_name: continue
            vals = []
            for row in data[1:]:
                if col_idx < len(row):
                    try: vals.append(float(str(row[col_idx]).replace("€","").replace(",",".").strip()))
                    except: pass
            if vals:
                lines.append(f"  {col_name}: total={sum(vals):.2f}, avg={sum(vals)/len(vals):.2f}, min={min(vals):.2f}, max={max(vals):.2f}")
        lines.append("  Primeras 15 filas:")
        for i, row in enumerate(data[1:16]):
            row_str = " | ".join(f"{headers[j] if j<len(headers) else j}={v}" for j,v in enumerate(row) if v)
            lines.append(f"    Fila {i+1}: {row_str}")
        if len(data) > 16:
            lines.append(f"    ... y {len(data)-16} más")
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════
#  EDITAR INTELIGENTE — AI genera una nueva hoja profesional
# ══════════════════════════════════════════════════════════════

async def ai_add_sheet(
    file_bytes: bytes,
    user_request: str,
    ai_client=None,
    model: str = "llama-3.3-70b-versatile"
) -> tuple[bytes, str]:
    """
    Reads the Excel, asks AI for a structured plan, builds a professional new sheet.
    Returns (excel_bytes, sheet_name_created).
    """
    from app.services.ai_service import ask_ai_for_json

    # 1. Parse existing workbook
    spreadsheet_text = _parse_excel(file_bytes)
    wb = load_workbook(io.BytesIO(file_bytes))

    # 2. Ask AI to generate the new sheet plan
    prompt = f"""Tienes este Excel con los siguientes datos:
{spreadsheet_text}

El usuario solicita: "{user_request}"

Crea un plan JSON para una nueva hoja profesional. Usa DATOS REALES del Excel.
Valores numéricos sin símbolos (123.45 no 123,45€). col_types: text|currency|percent|number|date.

{{
  "sheet_name": "Nombre (máx 28 chars)",
  "title": "Título principal",
  "subtitle": "Subtítulo",
  "sections": [
    {{"type": "header", "text": "Sección"}},
    {{"type": "table", "title": "Tabla", "headers": ["Col1","Col2","Col3"],
      "rows": [["dato","dato","dato"]], "has_totals": true,
      "col_types": ["text","currency","number"]}},
    {{"type": "kpi", "items": [{{"label": "KPI", "value": "123.45", "type": "currency"}}]}},
    {{"type": "text", "content": "Conclusión"}}
  ],
  "source_note": "FiscalIA"
}}"""

    raw = await ask_ai_for_json(
        prompt,
        system="Eres experto en Excel financiero. Responde SOLO con JSON válido, sin texto ni markdown."
    )
    plan = json.loads(raw)

    # 3. Build the professional sheet
    sheet_name = plan.get("sheet_name", "Análisis IA")[:28]
    # Avoid duplicate names
    existing = wb.sheetnames
    if sheet_name in existing:
        sheet_name = f"{sheet_name[:24]} (IA)"

    ws = wb.create_sheet(sheet_name)
    _build_ai_sheet(ws, plan)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), sheet_name


def _build_ai_sheet(ws, plan: dict):
    """Renders the AI-generated plan into a professional Excel sheet."""
    # ── Main title ──
    ws.merge_cells("A1:H1")
    ws["A1"] = plan.get("title", "Análisis IA")
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill = _hfill(NAVY)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    ws["A2"] = plan.get("subtitle", "")
    ws["A2"].font = Font(name="Arial", size=10, color="888888", italic=True)
    ws["A2"].fill = _hfill("F0EDE4")
    ws["A2"].alignment = _left()
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:H3")
    ws["A3"] = f"Generado por FiscalIA · {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A3"].font = Font(name="Arial", size=8, color="AAAAAA")
    ws["A3"].alignment = _left()
    ws.row_dimensions[3].height = 14

    current_row = 5
    sections = plan.get("sections", [])

    for section in sections:
        stype = section.get("type", "text")

        if stype == "header":
            ws.merge_cells(f"A{current_row}:H{current_row}")
            ws[f"A{current_row}"] = section.get("text", "")
            ws[f"A{current_row}"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
            ws[f"A{current_row}"].fill = _hfill(GOLD)
            ws[f"A{current_row}"].alignment = _left()
            ws.row_dimensions[current_row].height = 22
            current_row += 2

        elif stype == "table":
            headers = section.get("headers", [])
            rows    = section.get("rows", [])
            col_types = section.get("col_types", [])
            has_totals = section.get("has_totals", False)
            title  = section.get("title", "")

            if title:
                ws.merge_cells(f"A{current_row}:H{current_row}")
                ws[f"A{current_row}"] = title
                ws[f"A{current_row}"].font = Font(name="Arial", bold=True, size=10, color=BLUE_)
                ws[f"A{current_row}"].fill = _hfill(LBLUE)
                ws[f"A{current_row}"].alignment = _left()
                ws.row_dimensions[current_row].height = 18
                current_row += 1

            # Header row
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(current_row, col_idx, h)
                cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
                cell.fill = _hfill(BLUE_)
                cell.alignment = _center()
                cell.border = _border()
                ws.column_dimensions[get_column_letter(col_idx)].width = max(
                    ws.column_dimensions[get_column_letter(col_idx)].width or 0,
                    max(len(str(h)) + 4, 12)
                )
            ws.row_dimensions[current_row].height = 20
            header_row = current_row
            current_row += 1

            # Data rows
            data_start = current_row
            fill_alt = [_hfill(WHITE), _hfill(GREY_L)]
            for r_idx, row in enumerate(rows):
                fill = fill_alt[r_idx % 2]
                for col_idx, val in enumerate(row, 1):
                    ctype = col_types[col_idx-1] if col_idx-1 < len(col_types) else "text"
                    cell = ws.cell(current_row, col_idx)
                    cell.border = _border()
                    cell.fill = fill
                    cell.font = _font()

                    # Try to parse numeric values
                    if ctype in ("currency", "number", "percent"):
                        try:
                            # Check if it's a formula
                            if str(val).startswith("="):
                                cell.value = str(val)
                            else:
                                num = float(str(val).replace("€","").replace("%","").replace(",",".").strip())
                                cell.value = num
                                if ctype == "currency": cell.number_format = EUR_FMT; cell.alignment = _right()
                                elif ctype == "percent": cell.number_format = PCT_FMT; cell.alignment = _right()
                                else: cell.number_format = '#,##0.00'; cell.alignment = _right()
                        except:
                            cell.value = str(val); cell.alignment = _left()
                    elif ctype == "date":
                        cell.value = str(val); cell.alignment = _center()
                    else:
                        cell.value = str(val); cell.alignment = _left()

                ws.row_dimensions[current_row].height = 16
                current_row += 1

            # Totals row
            if has_totals and rows:
                ws.merge_cells(f"A{current_row}:A{current_row}")
                ws[f"A{current_row}"] = "TOTAL"
                ws[f"A{current_row}"].font = Font(name="Arial", bold=True, color=WHITE)
                ws[f"A{current_row}"].fill = _hfill(NAVY)
                ws[f"A{current_row}"].alignment = _right()
                ws[f"A{current_row}"].border = _border()

                for col_idx in range(2, len(headers) + 1):
                    ctype = col_types[col_idx-1] if col_idx-1 < len(col_types) else "text"
                    if ctype in ("currency", "number"):
                        col_letter = get_column_letter(col_idx)
                        cell = ws.cell(current_row, col_idx)
                        cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{current_row-1})"
                        cell.font = Font(name="Arial", bold=True, color=WHITE)
                        cell.fill = _hfill(NAVY)
                        cell.number_format = EUR_FMT
                        cell.alignment = _right()
                        cell.border = _border()
                    else:
                        cell = ws.cell(current_row, col_idx)
                        cell.fill = _hfill(NAVY)
                        cell.border = _border()

                ws.row_dimensions[current_row].height = 20
                current_row += 1

            current_row += 2

        elif stype == "kpi":
            items = section.get("items", [])
            # KPIs in a row, 2 per row max width
            items_per_row = min(4, len(items))
            per_col = 2  # cols per KPI

            for k_idx, item in enumerate(items):
                col_start = 1 + (k_idx % items_per_row) * per_col
                if k_idx > 0 and k_idx % items_per_row == 0:
                    current_row += 4

                label_cell = ws.cell(current_row, col_start)
                val_cell   = ws.cell(current_row + 1, col_start)

                label_cell.value = item.get("label", "")
                label_cell.font = Font(name="Arial", size=9, color="666666")
                label_cell.fill = _hfill(LBLUE)
                label_cell.alignment = _center()
                label_cell.border = _border()

                raw_val = str(item.get("value", ""))
                ktype = item.get("type", "text")
                try:
                    if not raw_val.startswith("="):
                        num = float(raw_val.replace("€","").replace("%","").replace(",",".").strip())
                        val_cell.value = num
                        if ktype == "currency": val_cell.number_format = EUR_FMT
                        elif ktype == "percent": val_cell.number_format = PCT_FMT
                    else:
                        val_cell.value = raw_val
                except:
                    val_cell.value = raw_val

                val_cell.font = Font(name="Arial", bold=True, size=14, color=NAVY)
                val_cell.fill = _hfill(WHITE)
                val_cell.alignment = _center()
                val_cell.border = _border()
                ws.row_dimensions[current_row].height = 16
                ws.row_dimensions[current_row+1].height = 28
                ws.column_dimensions[get_column_letter(col_start)].width = max(
                    ws.column_dimensions[get_column_letter(col_start)].width or 0, 18
                )

            current_row += 5

        elif stype == "text":
            ws.merge_cells(f"A{current_row}:H{current_row}")
            ws[f"A{current_row}"] = section.get("content", "")
            ws[f"A{current_row}"].font = Font(name="Arial", size=9, color="444444", italic=True)
            ws[f"A{current_row}"].fill = _hfill(GREY_L)
            ws[f"A{current_row}"].alignment = _left()
            ws.row_dimensions[current_row].height = 30
            current_row += 2

    # ── Source note ──
    current_row += 1
    ws.merge_cells(f"A{current_row}:H{current_row}")
    ws[f"A{current_row}"] = plan.get("source_note", "Fuente: FiscalIA")
    ws[f"A{current_row}"].font = Font(name="Arial", size=8, color="AAAAAA", italic=True)
    ws[f"A{current_row}"].alignment = _left()

    # ── Set default column widths ──
    for col in range(1, 9):
        letter = get_column_letter(col)
        if not ws.column_dimensions[letter].width or ws.column_dimensions[letter].width < 12:
            ws.column_dimensions[letter].width = 14

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"


# ══════════════════════════════════════════════════════════════
#  GENERAR INFORMES DESDE CERO
# ══════════════════════════════════════════════════════════════

def generate_invoice_report(invoices: list, year: int) -> bytes:
    wb = Workbook()
    _build_invoice_sheet(wb, invoices, year)
    _build_summary_sheet(wb, invoices, year)
    _build_quarterly_sheet(wb, invoices, year)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _build_invoice_sheet(wb, invoices, year):
    ws = wb.create_sheet("Facturas")
    ws.merge_cells("A1:K1")
    ws["A1"] = f"FiscalIA — Registro de Facturas {year}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws["A1"].fill = _hfill(NAVY)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = _font(color="888888", size=9)
    ws.row_dimensions[2].height = 14

    headers = ["#","Fecha","Tipo","Emisor","Concepto","Base Imp.","IVA%","Cuota IVA","Total","Cuenta PGC","Deducible"]
    widths  = [5, 11, 9, 20, 28, 13, 8, 12, 12, 12, 10]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(4, col, h)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=9)
        cell.fill = _hfill(GOLD)
        cell.alignment = _center()
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 18

    fill_in = _hfill(LGREEN)
    fill_ga = _hfill(LORANG)
    for i, inv in enumerate(invoices, 1):
        row = 4 + i
        ref = getattr(inv,'fecha',None) or getattr(inv,'created_at',None)
        fill = fill_in if getattr(inv,'tipo','')=='ingreso' else fill_ga
        vals = [
            i,
            ref.strftime("%d/%m/%Y") if ref else "",
            getattr(inv,'tipo','').upper(),
            getattr(inv,'emisor','') or "",
            getattr(inv,'concepto','') or "",
            float(getattr(inv,'base_imponible',0) or 0),
            float(getattr(inv,'tipo_iva',21) or 21),
            float(getattr(inv,'cuota_iva',0) or 0),
            float(getattr(inv,'total',0) or 0),
            getattr(inv,'categoria','') or "",
            f"{getattr(inv,'porcentaje_deduccion',0) or 0:.0f}%" if getattr(inv,'deducible',False) else "No",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row, col, val)
            cell.fill = fill; cell.border = _border(); cell.font = _font(size=9)
            if col in (6,8,9): cell.number_format = EUR_FMT; cell.alignment = _right()
            elif col == 7: cell.number_format = '0"%"'; cell.alignment = _center()
            elif col == 1: cell.alignment = _center()
            else: cell.alignment = _left()
        ws.row_dimensions[row].height = 15

    last = 4 + len(invoices)
    total_row = last + 2
    ws.merge_cells(f"A{total_row}:E{total_row}")
    ws[f"A{total_row}"] = "TOTALES"
    ws[f"A{total_row}"].font = Font(name="Arial", bold=True, color=WHITE)
    ws[f"A{total_row}"].fill = _hfill(NAVY)
    ws[f"A{total_row}"].alignment = _right()
    for col, cl in [(6,"F"),(8,"H"),(9,"I")]:
        cell = ws.cell(total_row, col, f"=SUM({cl}5:{cl}{last})")
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.fill = _hfill(NAVY)
        cell.number_format = EUR_FMT
        cell.alignment = _right()
        cell.border = _border()
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}4"
    ws.sheet_view.showGridLines = False


def _build_summary_sheet(wb, invoices, year):
    ws = wb.create_sheet("Resumen Fiscal")
    ingresos_list = [i for i in invoices if getattr(i,'tipo','')=='ingreso']
    gastos_list   = [i for i in invoices if getattr(i,'tipo','')=='gasto']
    total_ing = sum(float(getattr(i,'base_imponible',0) or 0) for i in ingresos_list)
    total_gas = sum(float(getattr(i,'base_imponible',0) or 0) for i in gastos_list)
    gas_ded   = sum(float(getattr(i,'base_imponible',0) or 0)*float(getattr(i,'porcentaje_deduccion',0) or 0)/100
                    for i in gastos_list if getattr(i,'deducible',False))
    iva_rep   = sum(float(getattr(i,'cuota_iva',0) or 0) for i in ingresos_list)
    iva_sop   = sum(float(getattr(i,'cuota_iva',0) or 0)*float(getattr(i,'porcentaje_deduccion',0) or 0)/100
                    for i in gastos_list if getattr(i,'deducible',False))

    ws.merge_cells("A1:D1")
    ws["A1"] = f"Resumen Fiscal {year} — FiscalIA"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=WHITE)
    ws["A1"].fill = _hfill(NAVY)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 28
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.sheet_view.showGridLines = False

    def section(start, title, rows_data):
        ws.merge_cells(f"A{start}:D{start}")
        ws[f"A{start}"] = title
        ws[f"A{start}"].font = Font(name="Arial", bold=True, color=WHITE, size=10)
        ws[f"A{start}"].fill = _hfill(GOLD)
        ws[f"A{start}"].alignment = _left()
        ws.row_dimensions[start].height = 18
        for off,(lbl,val,fmt,note) in enumerate(rows_data,1):
            r = start+off
            ws[f"A{r}"] = lbl
            ws[f"A{r}"].font = _font(size=10)
            ws[f"A{r}"].fill = _hfill(GREY_L if off%2 else WHITE)
            ws[f"A{r}"].border = _border(); ws[f"A{r}"].alignment = _left()
            ws[f"B{r}"] = val
            ws[f"B{r}"].number_format = fmt
            ws[f"B{r}"].font = Font(name="Arial", bold=True, size=10)
            ws[f"B{r}"].border = _border(); ws[f"B{r}"].alignment = _right()
            if note:
                ws[f"C{r}"] = note
                ws[f"C{r}"].font = Font(name="Arial", size=9, color="666666", italic=True)
                ws[f"C{r}"].alignment = _left()
            ws.row_dimensions[r].height = 16
        return start + len(rows_data) + 2

    r = section(3, "💰 INGRESOS Y GASTOS", [
        ("Ingresos totales (base imponible)", total_ing, EUR_FMT, f"{len(ingresos_list)} facturas de ingreso"),
        ("Gastos totales (base imponible)",   total_gas, EUR_FMT, f"{len(gastos_list)} facturas de gasto"),
        ("Gastos fiscalmente deducibles",     gas_ded,   EUR_FMT, "Aplicando % deducción asignado"),
        ("Beneficio neto",                    total_ing-total_gas, EUR_FMT, "Ingresos − Gastos"),
        ("Margen neto (%)",                   (total_ing-total_gas)/total_ing if total_ing else 0, PCT_FMT, ""),
    ])
    r = section(r, "🏛 IVA (Modelo 303)", [
        ("IVA repercutido cobrado a clientes", iva_rep,              EUR_FMT, "IVA en facturas de ingreso"),
        ("IVA soportado deducible",            iva_sop,              EUR_FMT, "IVA pagado en gastos deducibles"),
        ("IVA a ingresar a Hacienda",          max(0,iva_rep-iva_sop), EUR_FMT, "Positivo = pagar · Negativo = devolver"),
    ])
    base_irpf = total_ing - gas_ded
    irpf_est  = _calc_irpf(base_irpf)
    section(r, "📊 IRPF ESTIMADO (orientativo — depende del tipo de contribuyente)", [
        ("Base liquidable estimada",    base_irpf,         EUR_FMT, "Ingresos − Gastos deducibles"),
        ("IRPF estimado anual",         irpf_est,          EUR_FMT, "Tramos estatales 2026 (sin tramo autonómico)"),
        ("IRPF retenido est. (15%)",    total_ing*0.15,    EUR_FMT, "Estimación. Verificar retenciones reales"),
        ("Nota",                        "Ver asesor fiscal", '@', "El IRPF real depende del tipo de contribuyente, CCAA y deducciones personales"),
    ])


def _build_quarterly_sheet(wb, invoices, year):
    ws = wb.create_sheet("Trimestral")
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Desglose Trimestral {year} — FiscalIA"
    ws["A1"].font = Font(name="Arial", bold=True, size=12, color=WHITE)
    ws["A1"].fill = _hfill(NAVY)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 26
    ws.sheet_view.showGridLines = False

    headers = ["Trimestre","Ingresos (€)","Gastos (€)","Beneficio (€)","IVA a pagar (€)","M.130 est. (€)"]
    widths  = [14,16,16,16,18,18]
    for col,(h,w) in enumerate(zip(headers,widths),1):
        cell = ws.cell(3, col, h)
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = _hfill(GOLD)
        cell.alignment = _center(); cell.border = _border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    q_map = {1:"T1",2:"T1",3:"T1",4:"T2",5:"T2",6:"T2",7:"T3",8:"T3",9:"T3",10:"T4",11:"T4",12:"T4"}
    quarters = {"T1":[],"T2":[],"T3":[],"T4":[]}
    for inv in invoices:
        ref = getattr(inv,'fecha',None) or getattr(inv,'created_at',None)
        if ref and ref.year == year:
            quarters[q_map.get(ref.month,"T4")].append(inv)

    fills = [_hfill(WHITE),_hfill(GREY_L)]
    for r_idx,(q,q_invs) in enumerate(quarters.items(),4):
        ing  = sum(float(getattr(i,'base_imponible',0) or 0) for i in q_invs if getattr(i,'tipo','')=='ingreso')
        gas  = sum(float(getattr(i,'base_imponible',0) or 0) for i in q_invs if getattr(i,'tipo','')=='gasto')
        iva_r= sum(float(getattr(i,'cuota_iva',0) or 0) for i in q_invs if getattr(i,'tipo','')=='ingreso')
        iva_s= sum(float(getattr(i,'cuota_iva',0) or 0)*float(getattr(i,'porcentaje_deduccion',0) or 0)/100
                   for i in q_invs if getattr(i,'tipo','')=='gasto' and getattr(i,'deducible',False))
        fill = fills[r_idx%2]
        for col,val in enumerate([q,ing,gas,ing-gas,max(0,iva_r-iva_s),max(0,(ing-gas)*0.20)],1):
            cell = ws.cell(r_idx, col, val)
            cell.border = _border(); cell.fill = fill
            cell.font = _font(bold=(col==1))
            if col==1: cell.alignment = _center()
            else: cell.number_format = EUR_FMT; cell.alignment = _right()
            if col==4 and isinstance(val,(int,float)) and val<0:
                cell.font = Font(name="Arial", color=RED_, bold=False)
        ws.row_dimensions[r_idx].height = 18

    tr = 8
    ws.cell(tr,1,"TOTAL AÑO").font = Font(name="Arial", bold=True, color=WHITE)
    ws.cell(tr,1).fill = _hfill(NAVY); ws.cell(tr,1).alignment = _center(); ws.cell(tr,1).border = _border()
    for cl,cn in [("B",2),("C",3),("D",4),("E",5),("F",6)]:
        cell = ws.cell(tr, cn, f"=SUM({cl}4:{cl}7)")
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.fill = _hfill(NAVY); cell.number_format = EUR_FMT
        cell.alignment = _right(); cell.border = _border()
    ws.row_dimensions[tr].height = 20


def _calc_irpf(base: float) -> float:
    if base <= 0: return 0.0
    tramos = [(12450,.19),(7750,.24),(15000,.30),(24800,.37),(240000,.45)]
    imp, rest = 0.0, base
    for lim, tipo in tramos:
        if rest <= 0: break
        g = min(rest, lim); imp += g*tipo; rest -= g
    if rest > 0: imp += rest*0.47
    return round(imp, 2)


def generate_csv_export(invoices: list) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["#","Fecha","Tipo","Emisor","Concepto","Base Imponible",
                     "Tipo IVA%","Cuota IVA","Total","Cuenta PGC","Deducible","% Deducción"])
    for i, inv in enumerate(invoices, 1):
        ref = getattr(inv,'fecha',None) or getattr(inv,'created_at',None)
        writer.writerow([
            i, ref.strftime("%d/%m/%Y") if ref else "",
            getattr(inv,'tipo',''), getattr(inv,'emisor','') or '',
            getattr(inv,'concepto','') or '',
            getattr(inv,'base_imponible',0), getattr(inv,'tipo_iva',21),
            getattr(inv,'cuota_iva',0), getattr(inv,'total',0),
            getattr(inv,'categoria','') or '',
            'Sí' if getattr(inv,'deducible',False) else 'No',
            getattr(inv,'porcentaje_deduccion',0),
        ])
    return output.getvalue().encode("utf-8-sig")


def add_rows_to_excel(file_bytes: bytes, new_rows: list, sheet_name=None) -> bytes:
    """Add rows to existing Excel matching column headers."""
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    fill_new = _hfill("E8F4E8")
    for row_data in new_rows:
        next_row = ws.max_row + 1
        for col_idx, header in enumerate(headers, 1):
            if header and header in row_data:
                cell = ws.cell(next_row, col_idx, row_data[header])
                cell.border = _border(); cell.fill = fill_new; cell.font = _font()
                if isinstance(row_data[header], float):
                    cell.number_format = EUR_FMT; cell.alignment = _right()
                else:
                    cell.alignment = _left()
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
