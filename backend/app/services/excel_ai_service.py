"""
FiscalIA — IA especializada en Excel
Motor independiente para creación y modificación de hojas de cálculo.
Convierte lenguaje natural en hojas Excel profesionales.
"""
import io, json, re
from datetime import datetime
from typing import Optional
from groq import AsyncGroq
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.core.config import settings

client = AsyncGroq(api_key=settings.GROQ_API_KEY)
MODEL = "llama-3.3-70b-versatile"  # Smart model for Excel — needs precision

# ── Style constants ────────────────────────────────────────────
NAVY   = "0D1B2A"; GOLD = "C9A84C"; WHITE = "FFFFFF"
BLUE_  = "1F4E79"; LBLUE = "D6E4F0"; LGREEN = "E8F5E9"
LORANG = "FFF3E0"; GREY_L = "F5F5F5"; GREY_D = "CCCCCC"
RED_   = "C0392B"

def _fill(c): return PatternFill("solid", start_color=c, fgColor=c)
def _font(bold=False, color="000000", size=10): return Font(name="Arial", bold=bold, color=color, size=size)
def _border():
    s = Side(style="thin", color=GREY_D)
    return Border(left=s, right=s, top=s, bottom=s)
def _center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def _right():  return Alignment(horizontal="right",  vertical="center")
def _left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

EUR_FMT = '#,##0.00 €;(#,##0.00 €);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'


EXCEL_AI_SYSTEM = """Eres un experto en Microsoft Excel y hojas de cálculo financieras.
Tu única tarea es generar planes JSON estructurados para crear o modificar hojas de cálculo Excel.

REGLAS:
1. Responde SIEMPRE SOLO con JSON válido. Ningún texto adicional.
2. Usa fórmulas Excel reales cuando sea posible (=SUM, =AVERAGE, =IF, etc.)
3. Los valores numéricos van sin símbolos de moneda (1234.56 no "1.234,56€")
4. Diseña con lógica profesional: títulos, secciones, subtotales, totales finales
5. Adapta el diseño al contexto (financiero, tabla simple, KPIs, comparativa, etc.)
6. Si el usuario pide añadir datos a un Excel existente, analiza el contexto y adapta

TIPOS DE CELDAS disponibles:
- text: texto plano
- currency: formato €  
- percent: formato %
- number: número genérico
- date: fecha DD/MM/YYYY
- formula: fórmula Excel (empieza con =)
- header: celda de encabezado (estilo especial)"""


async def parse_user_intent(
    user_request: str,
    excel_context: Optional[str] = None,
) -> dict:
    """
    Uses AI to understand what the user wants and returns a structured plan.
    """
    context_section = ""
    if excel_context:
        context_section = f"\nCONTEXTO DEL EXCEL EXISTENTE:\n{excel_context}\n"

    prompt = f"""{context_section}
El usuario solicita: "{user_request}"

Genera un JSON con el plan completo para crear/modificar una hoja Excel.
El JSON debe tener esta estructura:

{{
  "action": "create_sheet | add_data | create_workbook",
  "sheet_name": "Nombre hoja (máx 28 chars)",
  "title": "Título principal",
  "subtitle": "Subtítulo opcional",
  "sections": [
    {{
      "type": "title_block",
      "title": "Texto del título de sección",
      "color": "NAVY | GOLD | BLUE"
    }},
    {{
      "type": "table",
      "title": "Nombre de la tabla",
      "headers": ["Columna 1", "Columna 2", "Columna 3"],
      "col_types": ["text", "currency", "percent"],
      "col_widths": [20, 15, 12],
      "rows": [
        ["Dato texto", 1234.56, 0.15],
        ["Otro dato", 2500.00, 0.22]
      ],
      "has_totals": true,
      "total_label": "TOTAL"
    }},
    {{
      "type": "kpi_row",
      "items": [
        {{"label": "Ingresos", "value": 50000, "type": "currency", "color": "green"}},
        {{"label": "Gastos", "value": 30000, "type": "currency", "color": "red"}},
        {{"label": "Margen", "value": 0.40, "type": "percent", "color": "blue"}}
      ]
    }},
    {{
      "type": "formula_row",
      "label": "Total acumulado",
      "formula": "=SUM(B5:B20)",
      "col": 2,
      "format": "currency"
    }},
    {{
      "type": "note",
      "text": "Nota: datos orientativos. Fuente: FiscalIA"
    }}
  ],
  "auto_filter": true,
  "freeze_header": true
}}

IMPORTANTE: Usa datos reales del contexto si los hay. Si no, crea datos de ejemplo coherentes con la petición."""

    response = await client.chat.completions.create(
        model=MODEL,
        messages=[
            {"role": "system", "content": EXCEL_AI_SYSTEM},
            {"role": "user", "content": prompt}
        ],
        max_tokens=3000,
        temperature=0.2,
    )

    raw = response.choices[0].message.content.strip()
    # Extract JSON
    if "```" in raw:
        parts = raw.split("```")
        for part in parts:
            p = part.strip()
            if p.startswith("json"): p = p[4:].strip()
            if p.startswith("{"): raw = p; break
    s = raw.find("{"); e = raw.rfind("}") + 1
    if s >= 0: raw = raw[s:e]
    return json.loads(raw)


def _apply_header_row(ws, row: int, headers: list, col_types: list, col_widths: list, header_color: str = BLUE_):
    """Apply styled header row."""
    for col_idx, (h, width) in enumerate(zip(headers, col_widths or [15]*len(headers)), 1):
        cell = ws.cell(row, col_idx, str(h))
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=10)
        cell.fill = _fill(header_color)
        cell.alignment = _center()
        cell.border = _border()
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            ws.column_dimensions[get_column_letter(col_idx)].width or 0, width
        )
    ws.row_dimensions[row].height = 20


def _apply_data_cell(cell, value, ctype: str, row_idx: int):
    """Apply value and format to a data cell."""
    fills = [_fill(WHITE), _fill(GREY_L)]
    cell.fill = fills[row_idx % 2]
    cell.border = _border()
    cell.font = _font()

    if ctype == "formula" or (isinstance(value, str) and value.startswith("=")):
        cell.value = str(value)
        cell.number_format = EUR_FMT
        cell.alignment = _right()
    elif ctype == "currency":
        try:
            cell.value = float(str(value).replace("€","").replace(",","."))
            cell.number_format = EUR_FMT
            cell.alignment = _right()
        except: cell.value = str(value); cell.alignment = _left()
    elif ctype == "percent":
        try:
            v = float(str(value).replace("%","").replace(",","."))
            cell.value = v/100 if v > 1 else v
            cell.number_format = PCT_FMT
            cell.alignment = _right()
        except: cell.value = str(value); cell.alignment = _left()
    elif ctype == "number":
        try:
            cell.value = float(str(value).replace(",","."))
            cell.number_format = '#,##0.00'
            cell.alignment = _right()
        except: cell.value = str(value); cell.alignment = _left()
    elif ctype == "date":
        cell.value = str(value)
        cell.alignment = _center()
    else:
        cell.value = str(value) if value is not None else ""
        cell.alignment = _left()


def build_excel_from_plan(plan: dict, existing_wb=None) -> tuple[bytes, str]:
    """
    Builds a professional Excel from the AI plan.
    Returns (excel_bytes, sheet_name).
    """
    wb = existing_wb or Workbook()
    if "Sheet" in wb.sheetnames and not existing_wb:
        del wb["Sheet"]

    sheet_name = plan.get("sheet_name", "FiscalIA")[:28]
    # Unique name
    base_name = sheet_name
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_name[:24]}_{counter}"
        counter += 1

    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    current_row = 1

    # ── Main title ──
    title = plan.get("title", sheet_name)
    subtitle = plan.get("subtitle", "")
    ws.merge_cells(f"A{current_row}:H{current_row}")
    ws[f"A{current_row}"] = title
    ws[f"A{current_row}"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws[f"A{current_row}"].fill = _fill(NAVY)
    ws[f"A{current_row}"].alignment = _center()
    ws.row_dimensions[current_row].height = 32
    current_row += 1

    if subtitle:
        ws.merge_cells(f"A{current_row}:H{current_row}")
        ws[f"A{current_row}"] = subtitle
        ws[f"A{current_row}"].font = Font(name="Arial", size=10, color="555555", italic=True)
        ws[f"A{current_row}"].fill = _fill("F0EDE4")
        ws[f"A{current_row}"].alignment = _left()
        ws.row_dimensions[current_row].height = 16
        current_row += 1

    ws.merge_cells(f"A{current_row}:H{current_row}")
    ws[f"A{current_row}"] = f"Generado por FiscalIA · {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws[f"A{current_row}"].font = Font(name="Arial", size=8, color="AAAAAA")
    ws.row_dimensions[current_row].height = 12
    current_row += 2

    header_colors = {"NAVY": NAVY, "GOLD": GOLD, "BLUE": BLUE_}

    for section in plan.get("sections", []):
        stype = section.get("type", "")

        # ── Title block ──
        if stype == "title_block":
            ws.merge_cells(f"A{current_row}:H{current_row}")
            ws[f"A{current_row}"] = section.get("title", "")
            color = header_colors.get(section.get("color","GOLD"), GOLD)
            ws[f"A{current_row}"].font = Font(name="Arial", bold=True, size=11, color=WHITE)
            ws[f"A{current_row}"].fill = _fill(color)
            ws[f"A{current_row}"].alignment = _left()
            ws.row_dimensions[current_row].height = 22
            current_row += 1

        # ── Table ──
        elif stype == "table":
            headers   = section.get("headers", [])
            col_types = section.get("col_types", ["text"] * len(headers))
            col_widths = section.get("col_widths", [16] * len(headers))
            rows      = section.get("rows", [])
            has_totals = section.get("has_totals", False)
            total_label = section.get("total_label", "TOTAL")

            tbl_title = section.get("title","")
            if tbl_title:
                ws.merge_cells(f"A{current_row}:H{current_row}")
                ws[f"A{current_row}"] = tbl_title
                ws[f"A{current_row}"].font = Font(name="Arial", bold=True, size=10, color=BLUE_)
                ws[f"A{current_row}"].fill = _fill(LBLUE)
                ws[f"A{current_row}"].alignment = _left()
                ws.row_dimensions[current_row].height = 16
                current_row += 1

            if headers:
                _apply_header_row(ws, current_row, headers, col_types, col_widths)
                header_row_num = current_row
                current_row += 1

            data_start = current_row
            for r_idx, row in enumerate(rows):
                for col_idx, (val, ctype) in enumerate(zip(row, col_types + ["text"]*(len(row)-len(col_types))), 1):
                    cell = ws.cell(current_row, col_idx)
                    _apply_data_cell(cell, val, ctype, r_idx)
                ws.row_dimensions[current_row].height = 16
                current_row += 1

            if has_totals and rows:
                # Total row
                ws.cell(current_row, 1, total_label).font = Font(name="Arial", bold=True, color=WHITE)
                ws.cell(current_row, 1).fill = _fill(NAVY)
                ws.cell(current_row, 1).alignment = _right()
                ws.cell(current_row, 1).border = _border()
                for col_idx, ctype in enumerate(col_types, 1):
                    if col_idx == 1: continue
                    cell = ws.cell(current_row, col_idx)
                    col_letter = get_column_letter(col_idx)
                    if ctype in ("currency", "number"):
                        cell.value = f"=SUM({col_letter}{data_start}:{col_letter}{current_row-1})"
                        cell.number_format = EUR_FMT if ctype == "currency" else '#,##0.00'
                        cell.alignment = _right()
                    cell.font = Font(name="Arial", bold=True, color=WHITE)
                    cell.fill = _fill(NAVY)
                    cell.border = _border()
                ws.row_dimensions[current_row].height = 20
                current_row += 1

            # Auto filter on header row
            if plan.get("auto_filter") and headers:
                ws.auto_filter.ref = f"A{header_row_num}:{get_column_letter(len(headers))}{header_row_num}"

            current_row += 1

        # ── KPI row ──
        elif stype == "kpi_row":
            items = section.get("items", [])
            kpi_colors = {"green": LGREEN, "red": LORANG, "blue": LBLUE, "gold": "FFF8E1", "grey": GREY_L}
            per_col = 2
            for k_idx, item in enumerate(items):
                col = 1 + k_idx * per_col
                color = kpi_colors.get(item.get("color","grey"), GREY_L)

                lbl_cell = ws.cell(current_row, col, item.get("label",""))
                lbl_cell.font = Font(name="Arial", size=9, color="555555")
                lbl_cell.fill = _fill(color)
                lbl_cell.alignment = _center()
                lbl_cell.border = _border()

                val_cell = ws.cell(current_row + 1, col)
                raw_val = item.get("value", 0)
                ktype = item.get("type", "text")
                try:
                    num = float(str(raw_val).replace("€","").replace("%","").replace(",","."))
                    if ktype == "percent":
                        val_cell.value = num/100 if num > 1 else num
                        val_cell.number_format = PCT_FMT
                    elif ktype == "currency":
                        val_cell.value = num
                        val_cell.number_format = EUR_FMT
                    else:
                        val_cell.value = num
                        val_cell.number_format = '#,##0.00'
                except:
                    val_cell.value = str(raw_val)

                val_cell.font = Font(name="Arial", bold=True, size=14, color=NAVY)
                val_cell.fill = _fill(WHITE)
                val_cell.alignment = _center()
                val_cell.border = _border()
                ws.row_dimensions[current_row].height = 16
                ws.row_dimensions[current_row+1].height = 28
                ws.column_dimensions[get_column_letter(col)].width = max(
                    ws.column_dimensions[get_column_letter(col)].width or 0, 18
                )

            current_row += 4

        # ── Formula row ──
        elif stype == "formula_row":
            label = section.get("label","")
            formula = section.get("formula","")
            col = section.get("col", 2)
            fmt = section.get("format","currency")

            ws.cell(current_row, 1, label).font = Font(name="Arial", bold=True)
            ws.cell(current_row, 1).alignment = _left()
            ws.cell(current_row, 1).border = _border()
            ws.cell(current_row, 1).fill = _fill(LBLUE)

            cell = ws.cell(current_row, col, formula)
            cell.number_format = EUR_FMT if fmt=="currency" else PCT_FMT if fmt=="percent" else '#,##0.00'
            cell.font = Font(name="Arial", bold=True, color=NAVY)
            cell.alignment = _right()
            cell.border = _border()
            cell.fill = _fill(LBLUE)
            ws.row_dimensions[current_row].height = 18
            current_row += 2

        # ── Note ──
        elif stype == "note":
            ws.merge_cells(f"A{current_row}:H{current_row}")
            ws[f"A{current_row}"] = section.get("text","")
            ws[f"A{current_row}"].font = Font(name="Arial", size=8, color="888888", italic=True)
            ws[f"A{current_row}"].fill = _fill(GREY_L)
            ws[f"A{current_row}"].alignment = _left()
            ws.row_dimensions[current_row].height = 20
            current_row += 2

    # Freeze header
    if plan.get("freeze_header"):
        ws.freeze_panes = "A5"

    # Default column widths
    for col in range(1, 9):
        letter = get_column_letter(col)
        if not ws.column_dimensions[letter].width or ws.column_dimensions[letter].width < 12:
            ws.column_dimensions[letter].width = 14

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), sheet_name


async def process_excel_request(
    user_request: str,
    file_bytes: Optional[bytes] = None,
    filename: Optional[str] = None,
) -> tuple[bytes, str, str]:
    """
    Main entry point for Excel AI.
    Returns (excel_bytes, sheet_name, explanation).
    """
    from app.services.spreadsheet_service import parse_spreadsheet_for_ai

    excel_context = None
    existing_wb = None

    if file_bytes and filename:
        try:
            excel_context = parse_spreadsheet_for_ai(file_bytes, filename)
            existing_wb = load_workbook(io.BytesIO(file_bytes))
        except Exception as e:
            excel_context = f"[Error leyendo el archivo: {str(e)}]"

    # AI generates the plan
    plan = await parse_user_intent(user_request, excel_context)

    # Build the Excel
    excel_bytes, sheet_name = build_excel_from_plan(plan, existing_wb)

    # Generate a short explanation
    explanation = f"He creado la hoja '{sheet_name}' con {len(plan.get('sections',[]))} secciones"
    if excel_context:
        explanation = f"He analizado tu Excel y añadido la hoja '{sheet_name}'"

    return excel_bytes, sheet_name, explanation


async def chat_with_excel_ai(
    message: str,
    conversation_history: Optional[list] = None,
) -> str:
    """
    Conversational interface for the Excel AI.
    Answers questions about Excel and helps plan spreadsheets.
    """
    messages = [
        {"role": "system", "content": """Eres un asistente especializado en Microsoft Excel y hojas de cálculo.
Ayudas a los usuarios a crear, modificar y entender Excel.
Cuando el usuario quiera crear o modificar un Excel, explícale qué vas a hacer y pídele que use el botón
'Generar Excel' para que se descargue el archivo.
Responde en español, de forma clara y práctica.
Si el usuario pregunta algo sobre Excel (fórmulas, formato, funciones), responde directamente.
Si quiere crear un Excel, describe brevemente lo que crearías y anímale a pulsar el botón."""},
    ]
    if conversation_history:
        messages.extend(conversation_history[-6:])
    messages.append({"role": "user", "content": message})

    response = await client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=messages,
        max_tokens=600,
        temperature=0.5,
    )
    return response.choices[0].message.content
